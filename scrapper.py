import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


def web_scrape(wb, web_address):
    # Send GET Request to Website
    print("Pinging", web_address)
    response = requests.get(web_address)

    # Check for valid response
    if response.status_code > 300 or response.status_code < 200:
        print("Possible Error, response code:", response.status_code)
        exit()
    else:
        print("Received Valid Response from website")

    # Look for tables
    table_class = input("Table class name (press enter for all tables): ").strip()  # Table's HTML class tag

    soup = BeautifulSoup(response.content, 'html.parser')
    if table_class == "\n":
        stat_table = soup.find_all('table')  # TODO: FIX no-class lookup
    else:
        stat_table = soup.find_all('table', class_=table_class)

    if len(table_class) == 0:
        print("Error: No specified tables found")
        return

    # Scrape tables, save data in excel sheet
    sheet_title = ""
    for x, table in enumerate(stat_table):
        sheet_title = workbook_title(table)
        ws = wb.create_sheet(sheet_title)
        for i, row in enumerate(table.find_all('tr')):
            for j, cell in enumerate(row.find_all('td')):  # place into excel here
                ws.cell(row=i + 1, column=j + 1, value=cell.text)
        print("Successful generated sheet", ws.title)


# Sets workbook title as
def workbook_title(table):
    try:
        title = table['id']
    except LookupError:
        try:
            title = table.parent['id']
        except LookupError:
            try:
                title = table['class'][0]
            except LookupError:
                title = "web scraped sheet"
    return title


def main():
    # Generate excel sheet
    workbook = Workbook()

    while True:
        web_address = input("Enter website URL (or enter to finish): ")
        if web_address == "":
            break
        elif web_address.find("http") == -1:
            continue
        web_scrape(workbook, web_address)

    if len(workbook.sheetnames) == 1:
        print("No file created")
        return

    while True:
        dest_filename = input("Destination file name: ").strip()  # Output .xlsx destination file
        if dest_filename != "":
            break
    workbook.remove(workbook.active)  # Clear blank default sheet
    workbook.save(filename=dest_filename + ".xlsx")
    print("Successful generated ", dest_filename, ".xls", sep="")
    quit()

input("\rWelcome to Web-Scrapper, press any key to continue to continue...")
print(end="\n\n")
main()
